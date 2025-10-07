import os
import csv
import openpyxl
import camelot  # pip install camelot-py[cv]
from collections import Counter
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# =========================
# CONFIGURAÇÕES
# =========================
PASTA_ATUAL = os.getcwd()
COLUNAS_EXCEL = ['I','K']  # Débito e Crédito
LINHA_INICIAL = 10
COR_DIVERGENCIA = "FFFF00"
fill = PatternFill(start_color=COR_DIVERGENCIA, end_color=COR_DIVERGENCIA, fill_type="solid")

# =========================
# FUNÇÕES AUXILIARES
# =========================
def limpar_valor(valor):
    if valor is None:
        return None
    v = str(valor).strip().replace('.', '').replace(',', '.')
    v = ''.join(c for c in v if c.isdigit() or c == '.')
    try:
        return float(v)
    except:
        return None

# Converter PDF em CSV
def pdf_para_csv(pdf_path):
    csv_path = pdf_path[:-4] + ".csv"
    try:
        tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
        all_data = []
        for t in tables:
            all_data.extend(t.df.values.tolist())
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(all_data)
        print(f"{pdf_path} convertido para {csv_path}")
        return csv_path
    except Exception as e:
        print(f"Erro ao processar {pdf_path}: {e}")
        return None

# Ler Excel
def ler_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    valores = []
    posicoes = []
    for i, row in enumerate(ws.iter_rows(min_row=LINHA_INICIAL), start=LINHA_INICIAL):
        for col in COLUNAS_EXCEL:
            idx = column_index_from_string(col)-1
            val = limpar_valor(row[idx].value)
            if val is not None:
                valores.append(val)
                posicoes.append((i, idx, val))
    return valores, wb, ws, posicoes

# Ler CSV
def ler_csv(path):
    valores = []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            for cell in row:
                val = limpar_valor(cell)
                if val is not None:
                    valores.append(val)
    return valores

# Comparar Excel X CSV
def comparar_excel_csv(excel_path, csv_path):
    valores_excel, wb, ws, posicoes = ler_excel(excel_path)
    valores_csv = ler_csv(csv_path)

    cont_excel = Counter(valores_excel)
    cont_csv = Counter(valores_csv)

    divergencias_excel = []
    divergencias_csv = []

    for val in cont_excel:
        diff = cont_excel[val] - cont_csv.get(val,0)
        if diff > 0:
            divergencias_excel.extend([val]*diff)
    for val in cont_csv:
        diff = cont_csv[val] - cont_excel.get(val,0)
        if diff > 0:
            divergencias_csv.extend([val]*diff)

    # Pintar células divergentes
    to_paint_excel = Counter(divergencias_excel)
    to_paint_csv = Counter(divergencias_csv)
    for row_idx, col_idx, valor in posicoes:
        if to_paint_excel.get(valor,0) > 0:
            ws.cell(row=row_idx, column=col_idx+1).fill = fill
            to_paint_excel[valor] -=1
        elif to_paint_csv.get(valor,0) > 0:
            ws.cell(row=row_idx, column=col_idx+1).fill = fill
            to_paint_csv[valor] -=1

    # Criar aba divergências
    if "Divergencias" in wb.sheetnames:
        ws_div = wb["Divergencias"]
    else:
        ws_div = wb.create_sheet("Divergencias")

    ws_div.append(["Valores no Excel e não no PDF"])
    for v in divergencias_excel:
        ws_div.append([v])
    ws_div.append([])
    ws_div.append(["Valores no PDF e não no Excel"])
    for v in divergencias_csv:
        ws_div.append([v])

    nome_saida = excel_path[:-5] + "_comparado.xlsx"
    wb.save(nome_saida)
    print(f"Comparação salva em {nome_saida}")

# =========================
# EXECUÇÃO AUTOMÁTICA PARA TODOS PARES
# =========================
pares = []
arquivos = os.listdir(PASTA_ATUAL)
for arq in arquivos:
    if arq.lower().endswith(".xlsx"):
        pdf_correspondente = arq[:-5] + ".pdf"
        if pdf_correspondente in arquivos:
            pares.append((arq, pdf_correspondente))

if not pares:
    print("Nenhum par Excel + PDF encontrado.")
    exit()

for excel_file, pdf_file in pares:
    print(f"\nProcessando par: {excel_file} + {pdf_file}")
    csv_file = pdf_para_csv(pdf_file)
    if csv_file:
        comparar_excel_csv(excel_file, csv_file)

print("\nTodos os pares processados com sucesso!")
